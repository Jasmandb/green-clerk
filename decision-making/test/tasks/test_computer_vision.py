# import unittest
from src.app_config import Classification, logging, Waste
import os
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color, colors

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import tensorflow as tf

logger = logging.getLogger(__name__)


class TestComputerVision:
    # def setUp(self):
    #     pass
    #
    # def tearDown(self):
    #     pass
    def __init__(self):
        self.type = {}
        self.test_images = '../confusion_matrix'
        self.model_file = '../../resources/new_mobile_model_org_aug.tflite'
        self.labels = ['Garbage', 'Glass', 'Metal', 'Paper', 'Plastic']
        self.input_mean = 0
        self.input_std = 255

    def run_computer_vision(self, image_location):
        interpreter = tf.lite.Interpreter(model_path=self.model_file)
        interpreter.allocate_tensors()

        input_details = interpreter.get_input_details()
        output_details = interpreter.get_output_details()

        # check the type of the input tensor
        floating_model = input_details[0]['dtype'] == np.float32

        # NxHxWxC, H:1, W:2
        height = input_details[0]['shape'][1]
        width = input_details[0]['shape'][2]
        img = Image.open(image_location).resize((width, height))

        # add N dim
        input_data = np.expand_dims(img, axis=0)

        if floating_model:
            input_data = (np.float32(input_data) - self.input_mean) / self.input_std

        interpreter.set_tensor(input_details[0]['index'], input_data)

        interpreter.invoke()

        output_data = interpreter.get_tensor(output_details[0]['index'])
        results = np.squeeze(output_data)

        top_k = results.argsort()[-5:][::-1]
        sorted_results = []
        for i in top_k:
            if floating_model:
                sorted_results.append('{:08.6f}: {}'.format(float(results[i]), self.labels[i]))
            else:
                sorted_results.append('{:08.6f}: {}'.format(float(results[i] / 255.0), self.labels[i]))

        return sorted_results, Classification[self.labels[top_k[0]].upper()]

    def test_create_confusion_matrix(self):
        images = os.listdir(self.test_images)
        wrong_naming_convention = False
        for image in images:
            if '_' not in image:
                logger.debug('image: {}'.format(image))
                wrong_naming_convention = True

        if wrong_naming_convention:
            exit('error please fix naming convention')
        images = [image.split('_') for image in images]
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Image Name'
        sheet['B1'] = 'Test Case Status'
        sheet['C1'] = Classification.GARBAGE
        sheet['D1'] = Classification.GLASS
        sheet['E1'] = Classification.METAL
        sheet['F1'] = Classification.PAPER
        sheet['G1'] = Classification.PLASTIC
        sheet['H1'] = 'Computer Vision Results'
        redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        greenFill = PatternFill('solid', colors.GREEN)
        yellowFill = PatternFill('solid', colors.YELLOW)
        confusion_counter = failed_test_cases = 0
        # confusion counters for garbage
        garbage_glass_confusion_counter = 0
        garbage_metal_confusion_counter = 0
        garbage_paper_confusion_counter = 0
        garbage_plastic_confusion_counter = 0
        # confusion counters for glass
        glass_metal_confusion_counter = 0
        glass_paper_confusion_counter = 0
        glass_plastic_confusion_counter = 0
        # confusion counters for metal
        metal_paper_confusion_counter = 0
        metal_plastic_confusion_counter = 0
        # confusion counters for paper
        paper_plastic_confusion_counter = 0

        logger.debug('Running Computer vision')
        for i, image in enumerate(images):
            try:
                image_location = '{}/{}_{}'.format(self.test_images, image[0], image[1])
                expected_result = Classification(image[0].upper())
                results, cv_result = self.run_computer_vision(image_location)
            except Exception:
                logger.debug('image: {}'.format(image))
                logger.debug(
                    'image_location: {}, expected_result: {}, results: {}, cv_result: {}'.format(image_location,
                                                                                                 expected_result,
                                                                                                 results, cv_result))
                logger.debug('image[0]: {}, image[1]: {}'.format(image[0], image[1]))

            # Input data to the excel sheet
            sheet['A{}'.format(i + 2)] = '{}_{}'.format(image[0], image[1])
            sheet['B{}'.format(i + 2)].fill = greenFill
            if expected_result == cv_result:
                sheet['B{}'.format(i + 2)].fill = greenFill
            else:
                sheet['B{}'.format(i + 2)].fill = redFill
                failed_test_cases += 1

            if expected_result == Classification.GARBAGE and expected_result == cv_result:
                sheet['C{}'.format(i + 2)].fill = greenFill
            elif expected_result == Classification.GLASS and expected_result == cv_result:
                sheet['D{}'.format(i + 2)].fill = greenFill
            elif expected_result == Classification.METAL and expected_result == cv_result:
                sheet['E{}'.format(i + 2)].fill = greenFill
            elif expected_result == Classification.PAPER and expected_result == cv_result:
                sheet['F{}'.format(i + 2)].fill = greenFill
            elif expected_result == Classification.PLASTIC and expected_result == cv_result:
                sheet['G{}'.format(i + 2)].fill = greenFill
            elif Classification.GARBAGE == cv_result:
                sheet['C{}'.format(i + 2)].fill = yellowFill
                confusion_counter += 1
                # Confusion with other class
                if expected_result == Classification.GLASS:
                    garbage_glass_confusion_counter += 1
                elif expected_result == Classification.METAL:
                    garbage_metal_confusion_counter += 1
                elif expected_result == Classification.PAPER:
                    garbage_paper_confusion_counter += 1
                elif expected_result == Classification.PLASTIC:
                    garbage_plastic_confusion_counter += 1
            elif Classification.GLASS == cv_result:
                sheet['D{}'.format(i + 2)].fill = yellowFill
                confusion_counter += 1
                # Confusion with other class
                if expected_result == Classification.GARBAGE:
                    garbage_glass_confusion_counter += 1
                elif expected_result == Classification.METAL:
                    glass_metal_confusion_counter += 1
                elif expected_result == Classification.PAPER:
                    glass_paper_confusion_counter += 1
                elif expected_result == Classification.PLASTIC:
                    glass_plastic_confusion_counter += 1
            elif Classification.METAL == cv_result:
                sheet['E{}'.format(i + 2)].fill = yellowFill
                confusion_counter += 1
                # Confusion with other class
                if expected_result == Classification.GARBAGE:
                    garbage_metal_confusion_counter += 1
                elif expected_result == Classification.GLASS:
                    glass_metal_confusion_counter += 1
                elif expected_result == Classification.PAPER:
                    metal_paper_confusion_counter += 1
                elif expected_result == Classification.PLASTIC:
                    metal_plastic_confusion_counter += 1
            elif Classification.PAPER == cv_result:
                sheet['F{}'.format(i + 2)].fill = yellowFill
                confusion_counter += 1
                if expected_result == Classification.PLASTIC:
                    paper_plastic_confusion_counter += 1
                elif expected_result == Classification.GARBAGE:
                    garbage_paper_confusion_counter += 1
                elif expected_result == Classification.GLASS:
                    glass_paper_confusion_counter += 1
                elif expected_result == Classification.METAL:
                    metal_paper_confusion_counter += 1
            elif Classification.PLASTIC == cv_result:
                sheet['G{}'.format(i + 2)].fill = yellowFill
                confusion_counter += 1
                if expected_result == Classification.PAPER:
                    paper_plastic_confusion_counter += 1
                elif expected_result == Classification.GARBAGE:
                    garbage_plastic_confusion_counter += 1
                elif expected_result == Classification.GLASS:
                    glass_plastic_confusion_counter += 1
                elif expected_result == Classification.METAL:
                    metal_plastic_confusion_counter += 1
            sheet['H{}'.format(i + 2)] = '{}'.format(results)

        # confusion counters for garbage
        sheet['I1'] = 'garbage_glass_confusion_counter'
        sheet['I2'] = garbage_glass_confusion_counter
        sheet['J1'] = 'garbage_metal_confusion_counter'
        sheet['J2'] = garbage_metal_confusion_counter
        sheet['I3'] = 'garbage_paper_confusion_counter'
        sheet['I4'] = garbage_paper_confusion_counter
        sheet['J3'] = 'garbage_plastic_confusion_counter'
        sheet['J4'] = garbage_plastic_confusion_counter
        # confusion counters for glass
        sheet['I5'] = 'glass_metal_confusion_counter'
        sheet['I6'] = glass_metal_confusion_counter
        sheet['J5'] = 'glass_paper_confusion_counter'
        sheet['J6'] = glass_paper_confusion_counter
        sheet['I7'] = 'glass_plastic_confusion_counter'
        sheet['I8'] = glass_plastic_confusion_counter
        # confusion counters for metal
        sheet['J7'] = 'metal_paper_confusion_counter'
        sheet['J8'] = metal_paper_confusion_counter
        sheet['I9'] = 'metal_plastic_confusion_counter'
        sheet['I10'] = metal_plastic_confusion_counter
        # confusion counters for paper
        sheet['J9'] = 'paper_plastic_confusion_counter'
        sheet['J10'] = paper_plastic_confusion_counter

        sheet['I11'] = 'confusion_counter'
        sheet['I12'] = confusion_counter

        sheet['J11'] = 'failed_test_cases'
        sheet['J12'] = failed_test_cases

        workbook.save(filename='{}/confusion_matrix_model_org_aug_results.xlsx'.format(self.test_images))


if __name__ == '__main__':
    TestComputerVision().test_create_confusion_matrix()
